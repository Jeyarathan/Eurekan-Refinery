"""Gulf Coast Excel parser — PIMS format to Eurekan models.

This module is a TRANSLATION LAYER. PIMS tags go in, Eurekan models come out.
No PIMS tag survives past the parser.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl

from eurekan.core.config import RefineryConfig, UnitConfig
from eurekan.core.crude import (
    CrudeAssay,
    CrudeLibrary,
    CutProperties,
    DistillationCut,
    US_GULF_COAST_630EP,
)
from eurekan.core.enums import DataSource, StreamDisposition, UnitType
from eurekan.core.product import Product, ProductSpec
from eurekan.core.stream import Stream
from eurekan.parsers.schema import SchemaValidationError, SheetSchema, validate_sheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PIMS → Eurekan translation maps (the ONLY place PIMS tags exist)
# ---------------------------------------------------------------------------

# Yield rows: tag → (eurekan_cut_name, sub_component_name_or_None)
PIMS_YIELD_MAP: dict[str, tuple[str, Optional[str]]] = {
    "VBALNC3": ("lpg", "propane"),
    "VBALIC4": ("lpg", "isobutane"),
    "VBALNC4": ("lpg", "n_butane"),
    "DBALLN1": ("light_naphtha", None),
    "DBALMN1": ("heavy_naphtha", None),
    "VBALKE1": ("kerosene", None),
    "VBALDS1": ("diesel", None),
    "VBALLV1": ("vgo", "lvgo"),
    "VBALHV1": ("vgo", "hvgo"),
    "VBALVR1": ("vacuum_residue", None),
}

# Property rows for each cut — tag → (eurekan_cut_name, property_name)
PIMS_PROPERTY_MAP: dict[str, tuple[str, str]] = {
    # Specific gravity
    "ISPGLN1": ("light_naphtha", "spg"),
    "ISPGMN1": ("heavy_naphtha", "spg"),
    "ISPGKE1": ("kerosene", "spg"),
    "ISPGDS1": ("diesel", "spg"),
    "ISPGLV1": ("vgo", "spg_lvgo"),
    "ISPGHV1": ("vgo", "spg_hvgo"),
    "ISPGVR1": ("vacuum_residue", "spg"),
    # API gravity
    "IAPIKE1": ("kerosene", "api"),
    "IAPIDS1": ("diesel", "api"),
    "IAPILV1": ("vgo", "api_lvgo"),
    "IAPIHV1": ("vgo", "api_hvgo"),
    "IAPIVR1": ("vacuum_residue", "api"),
    # Sulfur
    "ISULLN1": ("light_naphtha", "sulfur"),
    "ISULMN1": ("heavy_naphtha", "sulfur"),
    "ISULKE1": ("kerosene", "sulfur"),
    "ISULDS1": ("diesel", "sulfur"),
    "ISULLV1": ("vgo", "sulfur_lvgo"),
    "ISULHV1": ("vgo", "sulfur_hvgo"),
    "ISULVR1": ("vacuum_residue", "sulfur"),
    # RON (light naphtha only)
    "IRONLN1": ("light_naphtha", "ron"),
    # CCR (VGO sub-cuts and vacuum residue)
    "ICCNLV1": ("vgo", "ccr_lvgo"),
    "ICCNHV1": ("vgo", "ccr_hvgo"),
    "ICCNVR1": ("vacuum_residue", "ccr"),
    # Nickel
    "INIKLV1": ("vgo", "nickel_lvgo"),
    "INIKHV1": ("vgo", "nickel_hvgo"),
    "INIKVR1": ("vacuum_residue", "nickel"),
    # Vanadium
    "IVANLV1": ("vgo", "vanadium_lvgo"),
    "IVANHV1": ("vgo", "vanadium_hvgo"),
    "IVANVR1": ("vacuum_residue", "vanadium"),
    # Nitrogen (as naphthenic acids — N2A rows)
    "IN2ALN1": ("light_naphtha", "nitrogen"),
    "IN2AMN1": ("heavy_naphtha", "nitrogen"),
}

# Product tag → Eurekan name
PIMS_PRODUCT_MAP: dict[str, str] = {
    "CRG": "regular_gasoline",
    "CPR": "premium_gasoline",
    "RRG": "rfg_regular",
    "RPR": "rfg_premium",
    "ULS": "ulsd",
    "JET": "jet_fuel",
    "N2O": "no2_oil",
    "LSF": "low_sulfur_fuel_oil",
    "HSF": "high_sulfur_fuel_oil",
    "LPG": "lpg",
    "BUT": "mixed_butanes",
    "CKE": "coke",
    "LSR": "light_straight_run",
    "BEN": "benzene",
    "TOL": "toluene",
    "XYL": "xylenes",
    "LUB": "lube_base_stocks",
    "ATB": "a960",
    "SUP": "sulfur",
}

# Schemas for validation
ASSAYS_SCHEMA = SheetSchema(
    sheet_name="Assays",
    required_row_tags=[
        "VBALNC3", "VBALIC4", "VBALNC4",
        "DBALLN1", "DBALMN1",
        "VBALKE1", "VBALDS1",
        "VBALLV1", "VBALHV1", "VBALVR1",
    ],
    required_column_tags=["ARL"],
)

BUY_SCHEMA = SheetSchema(
    sheet_name="Buy",
    required_row_tags=["ARL"],
)

SELL_SCHEMA = SheetSchema(
    sheet_name="Sell",
    required_row_tags=["CRG", "JET", "ULS"],
)

BLNSPEC_SCHEMA = SheetSchema(
    sheet_name="Blnspec",
    required_row_tags=["NDON", "XRVI", "XSUL"],
    required_column_tags=["CRG"],
)

BLNMIX_SCHEMA = SheetSchema(
    sheet_name="Blnmix",
    required_row_tags=["LCN", "HCN", "NC4"],
    required_column_tags=["CRG"],
)

BLNNAPH_SCHEMA = SheetSchema(
    sheet_name="Blnnaph",
    required_row_tags=["NC4", "LCN", "HCN"],
)

CAPS_SCHEMA = SheetSchema(
    sheet_name="Caps",
    required_row_tags=["CAT1", "CCCU"],
)

PROCLIM_SCHEMA = SheetSchema(
    sheet_name="ProcLim",
    required_row_tags=["ZCVN", "ZRTT"],
)

# Spec tag → (eurekan_property_name, bound_type)
PIMS_SPEC_MAP: dict[str, tuple[str, str]] = {
    "NDON": ("road_octane", "min"),
    "XRVI": ("rvp_index", "max"),
    "XSUL": ("sulfur", "max"),
    "XBNZ": ("benzene", "max"),
    "XARO": ("aromatics", "max"),
    "XOLF": ("olefins", "max"),
}

# Component tag → Eurekan stream name
PIMS_COMPONENT_MAP: dict[str, str] = {
    "LCN": "fcc_light_naphtha",
    "HCN": "fcc_heavy_naphtha",
    "LN1": "cdu_light_naphtha",
    "LN2": "cdu_light_naphtha_2",
    "LN3": "cdu_light_naphtha_3",
    "NC4": "n_butane",
    "IC4": "isobutane",
    "RFT": "reformate",
    "ALK": "alkylate",
    "DIM": "dimate",
    "LKT": "treated_lt_coker_naphtha",
    "ISM": "isomerate",
    "P5S": "aru_pentanes",
    "RAF": "raffinate",
    "TOL": "toluene",
    "XYL": "xylenes",
    "C9A": "c9_aromatics",
    "SCN": "scanfinate",
    "HCL": "hc_light_naphtha",
    "ETR": "ethanol_rrg",
    "ETP": "ethanol_rpr",
    "NC3": "propane",
}

# Blnnaph property header → CutProperties field
PIMS_NAPH_PROP_MAP: dict[str, str] = {
    "RON": "ron",
    "MON": "mon",
    "RVI": "rvp",  # RVP index
    "SUL": "sulfur",
    "OLF": "olefins",
    "ARO": "aromatics",
    "BNZ": "benzene",
    "SPG": "spg",
}

# Caps tag → (eurekan_unit_id, UnitType, description)
PIMS_CAPS_MAP: dict[str, tuple[str, UnitType, str]] = {
    "CAT1": ("cdu_1", UnitType.CDU, "Crude Unit #1"),
    "CCCU": ("fcc_1", UnitType.FCC, "Cat Cracker"),
    "CLPR": ("reformer_1", UnitType.REFORMER, "Catalytic Reformer"),
    "CGHT": ("goht_1", UnitType.HYDROTREATER, "GO Hydrotreater"),
    "CGTU": ("scanfiner_1", UnitType.HYDROTREATER, "Scanfiner (FCC Naphtha HT)"),
    "CSFA": ("alky_1", UnitType.ALKYLATION, "Alkylation Unit"),
    "CKHT": ("kht_1", UnitType.HYDROTREATER, "Kero Hydrotreater"),
    "CDHT": ("dht_1", UnitType.HYDROTREATER, "Diesel Hydrotreater"),
    # Sprint 12: Vacuum unit + Delayed coker
    "CVT1": ("vacuum_1", UnitType.VACUUM, "Vacuum Distillation Unit"),
    "CDLC": ("coker_1", UnitType.COKER, "Delayed Coker"),
    # Sprint 13: Hydrocracker
    "CHCU": ("hcu_1", UnitType.HYDROCRACKER, "Hydrocracker"),
}

# ProcLim tag → (equipment_limit_key, description)
PIMS_PROCLIM_MAP: dict[str, str] = {
    "ZCVN": "fcc_conversion",
    "ZRTT": "fcc_riser_temp",
    "ZRGT": "fcc_regen_temp",
    "ZPHT": "fcc_preheat_temp",
    "ZAP1": "cdu1_api",
    "ZSU1": "cdu1_sulfur",
    "ZTA1": "cdu1_tan",
}

# Temperature ranges from the US Gulf Coast 630EP template, keyed by cut name
_CUT_TEMPS: dict[str, tuple[Optional[float], Optional[float], str]] = {}
for _cpd in US_GULF_COAST_630EP.cuts:
    _CUT_TEMPS[_cpd.name] = (_cpd.tbp_start_f, _cpd.tbp_end_f, _cpd.display_name)
# LPG has no TBP range in the template — add manually
_CUT_TEMPS["lpg"] = (None, None, "LPG (C3-C4)")


class GulfCoastParser:
    """Parser for the Gulf Coast PIMS-format Excel workbook."""

    def __init__(self, filepath: str | Path) -> None:
        self._filepath = Path(filepath)
        self._wb = openpyxl.load_workbook(self._filepath, data_only=True)

    @property
    def sheet_names(self) -> list[str]:
        return self._wb.sheetnames

    def explore_sheet(self, sheet_name: str, n_rows: int = 15) -> None:
        """Print the first *n_rows* of a sheet with row/column indices."""
        ws = self._wb[sheet_name]
        print(f"\n{'=' * 80}")
        print(f"Sheet: {sheet_name}  (rows={ws.max_row}, cols={ws.max_column})")
        print(f"{'=' * 80}")

        for row_idx, row in enumerate(
            ws.iter_rows(max_row=n_rows, values_only=False), start=1
        ):
            values = []
            for cell in row:
                v = cell.value
                if v is None:
                    values.append("")
                elif isinstance(v, float):
                    values.append(f"{v:.4f}")
                else:
                    values.append(str(v))
            truncated = [s[:12].ljust(12) for s in values]
            print(f"  Row {row_idx:3d}: {' | '.join(truncated)}")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _build_row_tag_index(self, ws) -> dict[str, int]:
        """Map column-A tags → row numbers (1-based)."""
        index: dict[str, int] = {}
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value is not None:
                tag = str(cell.value).strip()
                if not tag.startswith("*"):
                    index[tag] = cell.row
        return index

    def _build_col_header_index(self, ws, header_row: int) -> dict[str, int]:
        """Map header tags → column numbers (1-based) from a given row."""
        index: dict[str, int] = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                tag = str(cell.value).strip()
                if tag and tag != "TEXT":
                    index[tag] = cell.column
        return index

    def _find_header_row(self, ws, known_tag: str = "ARL") -> int:
        """Find the row containing crude/column tags (e.g. 'ARL')."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() == known_tag:
                    return cell.row
        raise SchemaValidationError(
            f"Could not find header row with tag '{known_tag}' in sheet"
        )

    def _cell_value(self, ws, row: int, col: int) -> Optional[float]:
        """Read a numeric cell value, returning None if empty or non-numeric."""
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    # ------------------------------------------------------------------
    # Task 1.7: Parse Assays sheet
    # ------------------------------------------------------------------

    def parse_assays(self) -> CrudeLibrary:
        """Parse the Assays sheet → CrudeLibrary with Eurekan-native names."""
        ws = self._wb["Assays"]

        # Validate schema
        issues = validate_sheet(ws, ASSAYS_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Assays sheet validation failed: {'; '.join(issues)}"
            )

        # Build indices
        header_row_num = self._find_header_row(ws, "ARL")
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        # Identify crude columns (skip TEXT column, skip empty)
        crude_tags = [
            tag for tag in col_index
            if tag not in ("TEXT", "***") and col_index[tag] > 2
        ]

        library = CrudeLibrary()

        for crude_tag in crude_tags:
            col = col_index[crude_tag]

            # --- Read whole-crude properties ---
            # API is in the row containing ZLIMAP1 (CDU-1 API) or from row 5 (*API Gravity)
            crude_api = self._cell_value(ws, row_index.get("ZLIMAP1", 0), col)
            if crude_api is None:
                # Fallback: read from row 5 (header section)
                crude_api = self._cell_value(ws, 5, col)
            crude_sulfur = self._cell_value(ws, row_index.get("ZLIMSU1", 0), col)
            if crude_sulfur is None:
                crude_sulfur = self._cell_value(ws, 9, col)
            crude_tan = self._cell_value(ws, row_index.get("ZLIMTA1", 0), col)

            if crude_api is None:
                crude_api = 30.0  # safe default
            if crude_sulfur is None:
                crude_sulfur = 1.0

            # --- Read yields ---
            # Accumulate sub-components into cuts
            cut_yields: dict[str, float] = {}
            for pims_tag, (cut_name, _sub) in PIMS_YIELD_MAP.items():
                if pims_tag not in row_index:
                    continue
                val = self._cell_value(ws, row_index[pims_tag], col)
                if val is not None:
                    cut_yields[cut_name] = cut_yields.get(cut_name, 0.0) + val

            # --- Read cut properties ---
            cut_props_raw: dict[str, dict[str, float]] = {}
            for pims_tag, (cut_name, prop_name) in PIMS_PROPERTY_MAP.items():
                if pims_tag not in row_index:
                    continue
                val = self._cell_value(ws, row_index[pims_tag], col)
                if val is not None:
                    cut_props_raw.setdefault(cut_name, {})[prop_name] = val

            # --- Build DistillationCut objects ---
            cuts: list[DistillationCut] = []
            for cut_name in [
                "lpg", "light_naphtha", "heavy_naphtha",
                "kerosene", "diesel", "vgo", "vacuum_residue",
            ]:
                vol_yield = cut_yields.get(cut_name, 0.0)
                if vol_yield == 0.0 and cut_name == "lpg":
                    continue  # Skip LPG if no yields found

                tbp_start, tbp_end, display = _CUT_TEMPS.get(
                    cut_name, (None, None, cut_name)
                )

                # Assemble CutProperties
                raw = cut_props_raw.get(cut_name, {})
                props = self._build_cut_properties(cut_name, raw)

                cuts.append(
                    DistillationCut(
                        name=cut_name,
                        display_name=display,
                        tbp_start_f=tbp_start,
                        tbp_end_f=tbp_end,
                        vol_yield=vol_yield,
                        properties=props,
                        source=DataSource.IMPORTED,
                        confidence=0.95,
                    )
                )

            # Read crude full name from row 2
            name_val = ws.cell(row=2, column=col).value
            crude_name = str(name_val).strip() if name_val else crude_tag

            assay = CrudeAssay(
                crude_id=crude_tag,
                name=crude_name,
                origin=None,
                api=crude_api,
                sulfur=crude_sulfur,
                tan=crude_tan,
                cuts=cuts,
            )
            library.add(assay)

        return library

    def _build_cut_properties(
        self, cut_name: str, raw: dict[str, float]
    ) -> CutProperties:
        """Build CutProperties from raw property dict, handling VGO aggregation."""
        if cut_name == "vgo":
            # Weighted average of LVGO and HVGO properties
            # Use simple average as approximation (true weighting needs volumes)
            api = self._avg(raw.get("api_lvgo"), raw.get("api_hvgo"))
            sulfur = self._avg(raw.get("sulfur_lvgo"), raw.get("sulfur_hvgo"))
            spg = self._avg(raw.get("spg_lvgo"), raw.get("spg_hvgo"))
            ccr = self._avg(raw.get("ccr_lvgo"), raw.get("ccr_hvgo"))
            nickel = self._avg(raw.get("nickel_lvgo"), raw.get("nickel_hvgo"))
            vanadium = self._avg(raw.get("vanadium_lvgo"), raw.get("vanadium_hvgo"))
            return CutProperties(
                api=api, sulfur=sulfur, spg=spg,
                ccr=ccr, nickel=nickel, vanadium=vanadium,
            )
        else:
            return CutProperties(
                api=raw.get("api"),
                sulfur=raw.get("sulfur"),
                spg=raw.get("spg"),
                ron=raw.get("ron"),
                nitrogen=raw.get("nitrogen"),
                ccr=raw.get("ccr"),
                nickel=raw.get("nickel"),
                vanadium=raw.get("vanadium"),
            )

    @staticmethod
    def _avg(a: Optional[float], b: Optional[float]) -> Optional[float]:
        """Average two optional values."""
        if a is not None and b is not None:
            return (a + b) / 2.0
        return a if a is not None else b

    # ------------------------------------------------------------------
    # Task 1.8: Parse Buy sheet
    # ------------------------------------------------------------------

    def parse_buy(self, library: CrudeLibrary) -> CrudeLibrary:
        """Parse Buy sheet and update CrudeAssay objects with price/availability."""
        ws = self._wb["Buy"]

        issues = validate_sheet(ws, BUY_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Buy sheet validation failed: {'; '.join(issues)}"
            )

        # Find header row by searching for 'COST' or 'TEXT'
        header_row_num = 3  # Known from exploration
        col_index = self._build_col_header_index(ws, header_row_num)

        cost_col = col_index.get("COST")
        min_col = col_index.get("MIN")
        max_col = col_index.get("MAX")
        api_col = col_index.get("API")
        sul_col = col_index.get("!SUL")

        row_index = self._build_row_tag_index(ws)

        for crude_tag, row_num in row_index.items():
            assay = library.get(crude_tag)
            if assay is None:
                continue

            price = self._cell_value(ws, row_num, cost_col) if cost_col else None
            min_rate = self._cell_value(ws, row_num, min_col) if min_col else None
            max_rate = self._cell_value(ws, row_num, max_col) if max_col else None

            if price is not None:
                assay.price = price
            if max_rate is not None:
                # Convert from '000 BPD → BPD
                assay.max_rate = max_rate * 1000.0
            if min_rate is not None:
                assay.min_rate = min_rate * 1000.0

        return library

    # ------------------------------------------------------------------
    # Task 1.8: Parse Sell sheet
    # ------------------------------------------------------------------

    def parse_sell(self) -> dict[str, Product]:
        """Parse Sell sheet → dict of Products keyed by Eurekan names."""
        ws = self._wb["Sell"]

        issues = validate_sheet(ws, SELL_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Sell sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)

        price_col = col_index.get("PRICE")
        min_col = col_index.get("MIN")
        max_col = col_index.get("MAX")

        row_index = self._build_row_tag_index(ws)
        products: dict[str, Product] = {}

        for pims_tag, row_num in row_index.items():
            eurekan_name = PIMS_PRODUCT_MAP.get(pims_tag)
            if eurekan_name is None:
                continue  # Skip products we don't map

            price = self._cell_value(ws, row_num, price_col) if price_col else None
            min_demand = self._cell_value(ws, row_num, min_col) if min_col else None
            max_demand = self._cell_value(ws, row_num, max_col) if max_col else None

            # Read display name from column B
            name_val = ws.cell(row=row_num, column=2).value
            display_name = str(name_val).strip() if name_val else eurekan_name

            products[eurekan_name] = Product(
                product_id=eurekan_name,
                name=display_name,
                price=price if price is not None else 0.0,
                min_demand=(min_demand * 1000.0) if min_demand is not None else 0.0,
                max_demand=(max_demand * 1000.0) if max_demand is not None else None,
            )

        return products

    # ------------------------------------------------------------------
    # Task 1.9: Parse Blnspec sheet
    # ------------------------------------------------------------------

    def parse_blnspec(self, products: dict[str, Product]) -> dict[str, Product]:
        """Parse Blnspec sheet and add specs to Product objects."""
        ws = self._wb["Blnspec"]

        issues = validate_sheet(ws, BLNSPEC_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Blnspec sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        # For each product column, extract specs
        for pims_prod_tag, prod_col in col_index.items():
            eurekan_name = PIMS_PRODUCT_MAP.get(pims_prod_tag)
            if eurekan_name is None or eurekan_name not in products:
                continue

            product = products[eurekan_name]
            specs: list[ProductSpec] = []

            for pims_spec_tag, (eurekan_prop, bound_type) in PIMS_SPEC_MAP.items():
                if pims_spec_tag not in row_index:
                    continue
                val = self._cell_value(ws, row_index[pims_spec_tag], prod_col)
                if val is not None:
                    if bound_type == "min":
                        specs.append(ProductSpec(spec_name=eurekan_prop, min_value=val))
                    else:
                        specs.append(ProductSpec(spec_name=eurekan_prop, max_value=val))

            if specs:
                product.specs = specs

        return products

    # ------------------------------------------------------------------
    # Task 1.9: Parse Blnmix sheet
    # ------------------------------------------------------------------

    def parse_blnmix(self, products: dict[str, Product]) -> dict[str, Product]:
        """Parse Blnmix sheet and set allowed_components on Product objects."""
        ws = self._wb["Blnmix"]

        issues = validate_sheet(ws, BLNMIX_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Blnmix sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        for pims_prod_tag, prod_col in col_index.items():
            eurekan_name = PIMS_PRODUCT_MAP.get(pims_prod_tag)
            if eurekan_name is None or eurekan_name not in products:
                continue

            product = products[eurekan_name]
            components: list[str] = []

            for pims_comp_tag, comp_row in row_index.items():
                val = self._cell_value(ws, comp_row, prod_col)
                if val is not None and val > 0:
                    eurekan_comp = PIMS_COMPONENT_MAP.get(pims_comp_tag)
                    if eurekan_comp is not None:
                        components.append(eurekan_comp)

            if components:
                product.allowed_components = components

        return products

    # ------------------------------------------------------------------
    # Task 1.9: Parse Blnnaph sheet
    # ------------------------------------------------------------------

    def parse_blnnaph(self) -> dict[str, CutProperties]:
        """Parse Blnnaph sheet → component properties keyed by Eurekan name."""
        ws = self._wb["Blnnaph"]

        issues = validate_sheet(ws, BLNNAPH_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Blnnaph sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        result: dict[str, CutProperties] = {}

        for pims_comp_tag, comp_row in row_index.items():
            eurekan_name = PIMS_COMPONENT_MAP.get(pims_comp_tag)
            if eurekan_name is None:
                continue

            props: dict[str, float] = {}
            for pims_prop_tag, eurekan_prop in PIMS_NAPH_PROP_MAP.items():
                prop_col = col_index.get(pims_prop_tag)
                if prop_col is None:
                    continue
                val = self._cell_value(ws, comp_row, prop_col)
                if val is not None:
                    props[eurekan_prop] = val

            result[eurekan_name] = CutProperties(**props)

        return result

    # ------------------------------------------------------------------
    # Task 1.10: Parse Caps sheet
    # ------------------------------------------------------------------

    def parse_caps(self) -> dict[str, UnitConfig]:
        """Parse Caps sheet → dict of UnitConfig objects."""
        ws = self._wb["Caps"]

        issues = validate_sheet(ws, CAPS_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"Caps sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        min_col = col_index.get("MIN")
        max_col = col_index.get("MAX")

        units: dict[str, UnitConfig] = {}

        for pims_tag, (unit_id, unit_type, desc) in PIMS_CAPS_MAP.items():
            if pims_tag not in row_index:
                continue

            row_num = row_index[pims_tag]
            min_val = self._cell_value(ws, row_num, min_col) if min_col else None
            max_val = self._cell_value(ws, row_num, max_col) if max_col else None

            # Caps values are in '000 BPD
            capacity = (max_val * 1000.0) if max_val is not None else 0.0
            min_tp = (min_val * 1000.0) if min_val is not None else 0.0

            units[unit_id] = UnitConfig(
                unit_id=unit_id,
                unit_type=unit_type,
                capacity=capacity,
                min_throughput=min_tp,
                source=DataSource.IMPORTED,
            )

        return units

    # ------------------------------------------------------------------
    # Task 1.10: Parse ProcLim sheet
    # ------------------------------------------------------------------

    def parse_proclim(
        self, units: dict[str, UnitConfig]
    ) -> dict[str, UnitConfig]:
        """Parse ProcLim sheet and add equipment_limits to UnitConfig objects."""
        ws = self._wb["ProcLim"]

        issues = validate_sheet(ws, PROCLIM_SCHEMA)
        if issues:
            raise SchemaValidationError(
                f"ProcLim sheet validation failed: {'; '.join(issues)}"
            )

        header_row_num = 3
        col_index = self._build_col_header_index(ws, header_row_num)
        row_index = self._build_row_tag_index(ws)

        min_col = col_index.get("MIN")
        max_col = col_index.get("MAX")

        # FCC limits go into fcc_1
        fcc = units.get("fcc_1")
        cdu = units.get("cdu_1")

        for pims_tag, limit_key in PIMS_PROCLIM_MAP.items():
            if pims_tag not in row_index:
                continue

            row_num = row_index[pims_tag]
            min_val = self._cell_value(ws, row_num, min_col) if min_col else None
            max_val = self._cell_value(ws, row_num, max_col) if max_col else None

            # Determine which unit this limit belongs to
            target = None
            if limit_key.startswith("fcc_") and fcc is not None:
                target = fcc
            elif limit_key.startswith("cdu1_") and cdu is not None:
                target = cdu

            if target is not None:
                if min_val is not None:
                    target.equipment_limits[f"{limit_key}_min"] = min_val
                if max_val is not None:
                    target.equipment_limits[f"{limit_key}_max"] = max_val

        return units

    # ------------------------------------------------------------------
    # Task 1.11: Assemble RefineryConfig
    # ------------------------------------------------------------------

    def parse(self) -> RefineryConfig:
        """Parse the entire Gulf Coast workbook → complete RefineryConfig."""
        # 1. Assays → CrudeLibrary
        library = self.parse_assays()

        # 2. Buy → update prices/availability
        self.parse_buy(library)

        # 3. Sell → products with prices
        products = self.parse_sell()

        # 4. Blnspec → add specs to products
        self.parse_blnspec(products)

        # 5. Blnmix → add allowed_components
        self.parse_blnmix(products)

        # 6. Blnnaph → component blend properties (stored separately)
        self.parse_blnnaph()

        # 7. Caps → UnitConfig for CDU and FCC
        units = self.parse_caps()

        # 8. ProcLim → equipment limits
        self.parse_proclim(units)

        # 9. Create Stream objects for CDU cuts and FCC products
        streams: dict[str, Stream] = {}

        # CDU output streams — one per cut
        cdu_cut_dispositions: dict[str, list[StreamDisposition]] = {
            "lpg": [StreamDisposition.SELL],
            "light_naphtha": [StreamDisposition.BLEND, StreamDisposition.SELL],
            "heavy_naphtha": [StreamDisposition.BLEND, StreamDisposition.SELL],
            "kerosene": [StreamDisposition.BLEND, StreamDisposition.SELL],
            "diesel": [StreamDisposition.BLEND, StreamDisposition.SELL],
            "vgo": [StreamDisposition.FCC_FEED, StreamDisposition.SELL, StreamDisposition.FUEL_OIL],
            "vacuum_residue": [StreamDisposition.FUEL_OIL, StreamDisposition.SELL],
        }
        for cut_name, dispositions in cdu_cut_dispositions.items():
            sid = f"cdu_{cut_name}"
            streams[sid] = Stream(
                stream_id=sid,
                source_unit="cdu_1",
                stream_type=cut_name,
                possible_dispositions=dispositions,
            )

        # FCC product streams
        fcc_streams: dict[str, list[StreamDisposition]] = {
            "fcc_light_naphtha": [StreamDisposition.BLEND],
            "fcc_heavy_naphtha": [StreamDisposition.BLEND, StreamDisposition.FUEL_OIL],
            "fcc_lco": [StreamDisposition.BLEND, StreamDisposition.FUEL_OIL],
            "fcc_slurry": [StreamDisposition.FUEL_OIL],
            "fcc_gas": [StreamDisposition.INTERNAL],
            "fcc_coke": [StreamDisposition.INTERNAL],
        }
        for stream_name, dispositions in fcc_streams.items():
            streams[stream_name] = Stream(
                stream_id=stream_name,
                source_unit="fcc_1",
                stream_type=stream_name,
                possible_dispositions=dispositions,
            )

        # 10. Assemble and return RefineryConfig
        return RefineryConfig(
            name="Gulf Coast Refinery",
            units=units,
            crude_library=library,
            products=products,
            streams=streams,
            cut_point_template=US_GULF_COAST_630EP,
        )
